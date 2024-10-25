
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file


def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
    salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['salario_empleado'])
    # convertir salario a INT
    salario_entero = int(salario_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_empleados (nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado, foto_empleado, salario_empleado) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_empleado'], dataForm['apellido_empleado'], dataForm['sexo_empleado'],
                           dataForm['telefono_empleado'], dataForm['email_empleado'], dataForm['profesion_empleado'], result_foto_perfil, salario_entero)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.foto_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_lista_empleadosBD: {e}")
        return None


# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.profesion_empleado,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None


# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.salario_empleado,
                        e.email_empleado,
                        e.telefono_empleado,
                        e.profesion_empleado,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido", "Sexo",
                     "Telefono", "Email", "Profesión", "Salario", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        sexo_empleado = registro['sexo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        profesion_empleado = registro['profesion_empleado']
        salario_empleado = registro['salario_empleado']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado,
                     salario_empleado, fecha_registro))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 7  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.salario_empleado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.sexo_empleado,
                            e.telefono_empleado,
                            e.email_empleado,
                            e.profesion_empleado,
                            e.salario_empleado,
                            e.foto_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_empleado = data.form['nombre_empleado']
                apellido_empleado = data.form['apellido_empleado']
                sexo_empleado = data.form['sexo_empleado']
                telefono_empleado = data.form['telefono_empleado']
                email_empleado = data.form['email_empleado']
                profesion_empleado = data.form['profesion_empleado']

                salario_sin_puntos = re.sub(
                    '[^0-9]+', '', data.form['salario_empleado'])
                salario_empleado = int(salario_sin_puntos)
                id_empleado = data.form['id_empleado']

                if data.files['foto_empleado']:
                    file = data.files['foto_empleado']
                    fotoForm = procesar_imagen_perfil(file)

                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            salario_empleado = %s,
                            foto_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado,
                              salario_empleado, fotoForm, id_empleado)
                else:
                    querySQL = """
                        UPDATE tbl_empleados
                        SET 
                            nombre_empleado = %s,
                            apellido_empleado = %s,
                            sexo_empleado = %s,
                            telefono_empleado = %s,
                            email_empleado = %s,
                            profesion_empleado = %s,
                            salario_empleado = %s
                        WHERE id_empleado = %s
                    """
                    values = (nombre_empleado, apellido_empleado, sexo_empleado,
                              telefono_empleado, email_empleado, profesion_empleado,
                              salario_empleado, id_empleado)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar uEmpleado
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
    
# Procesar Tarea
def procesar_form_tarea(dataForm):
    # Mapeo de estado y prioridad
    estado_mapeo = {
        "Por hacer": 1,
        "En progreso": 2,
        "Terminado": 3
    }
    
    prioridad_mapeo = {
        "Alta": 1,
        "Media": 2,
        "Baja": 3
    }
    
    # Obtener valores del formulario
    titulo = dataForm['titulo']
    descripcion = dataForm['descripcion']
    proyecto = dataForm['proyecto']
    estado = estado_mapeo.get(dataForm['estado'], 0)  # 0 si no coincide con ningún valor
    prioridad = prioridad_mapeo.get(dataForm['prioridad'], 0)
    fecha_inicio = dataForm['fecha_inicio']
    fecha_vencimiento = dataForm['fecha_vencimiento']
    id_empleado_asignado = dataForm['id_empleado_asignado']
    
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = """
                INSERT INTO tbl_tareas (titulo, descripcion, proyecto, estado, prioridad, fecha_inicio, fecha_vencimiento, id_empleado_asignado) 
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """
                valores = (titulo, descripcion, proyecto, estado, prioridad, fecha_inicio, fecha_vencimiento, id_empleado_asignado)
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                return cursor.rowcount
    except Exception as e:
        return f'Error al agregar la tarea: {str(e)}'
    
# procesar ticket
def procesar_form_ticket(dataForm):
    # Asegurarse de que los datos necesarios están presentes
    try:
        # Conexión a la base de datos y ejecución de la consulta
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # Consulta SQL para insertar el ticket en la tabla tbl_tickets
                sql = """
                    INSERT INTO tbl_tickets 
                    (titulo_ticket, descripcion_ticket, tipo_ticket, id_user, id_empleado_asignado) 
                    VALUES (%s, %s, %s, %s, %s)
                """
                
                # Obtener los valores del formulario
                valores = (
                    dataForm['titulo_ticket'],
                    dataForm['descripcion_ticket'],
                    int(dataForm['tipo_ticket']),
                    int(dataForm['id_user']),
                    int(dataForm['id_empleado_asignado'])
                )
                
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                
                # Retornar éxito si se inserta correctamente
                return cursor.rowcount > 0

    except Exception as e:
        # Retornar el error para depuración
        print(f"Error en procesar_form_ticket: {e}")
        return None

# procesar venta
def procesar_form_venta(datos_venta):
    try:
        # Obtener y validar datos
        id_cliente = int(datos_venta.get('id_cliente'))
        proyecto = datos_venta.get('proyecto')
        empresa = datos_venta.get('empresa')
        fecha_cobro = datos_venta.get('fecha_cobro')
        fecha_venta_vencimiento = datos_venta.get('fecha_venta_vencimiento')
        
        # Empaquetar los datos en una tupla
        return (id_cliente, proyecto, empresa, fecha_cobro, fecha_venta_vencimiento)
    
    except Exception as e:
        # Retornar un mensaje de error si ocurre algún problema
        return f"Error al procesar los datos de la venta: {e}"

def procesar_form_contacto(dataForm):
    # Verificar el valor del campo 'sexo_empleado'
    sexo_empleado_valor = 0  # Valor predeterminado para "Seleccione"
    
    if dataForm['sexo_empleado'] == "Masculino":
        sexo_empleado_valor = 1
    elif dataForm['sexo_empleado'] == "Femenino":
        sexo_empleado_valor = 2
    
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # Validar que el id_miembro_responsable exista en la tabla 'miembros'
                cursor.execute("SELECT COUNT(*) as count FROM tbl_empleados WHERE id_empleado = %s", (dataForm['id_miembro_responsable'],))
                resultado = cursor.fetchone()
                
                if resultado['count'] == 0:
                    # Si el id_miembro_responsable no existe, retornar un error
                    print(f"El miembro responsable con ID {dataForm['id_miembro_responsable']} no existe.")
                    return None

                # Consulta SQL para insertar el contacto
                sql = """
                    INSERT INTO tbl_contactos 
                    (nombre, apellido, email, teléfono, empresa, sexo_empleado, propietario, foto_empleado, id_miembro_responsable) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                
                # Obtener los valores del formulario
                valores = (
                    dataForm['nombre'],
                    dataForm['apellido'],
                    dataForm['email'],
                    dataForm['teléfono'],
                    dataForm['empresa'],
                    sexo_empleado_valor,
                    dataForm['propietario'],
                    dataForm['foto_empleado'],
                    int(dataForm['id_miembro_responsable'])
                )
                
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                
                # Retornar éxito si se inserta correctamente
                return cursor.rowcount > 0

    except Exception as e:
        print(f"Error en procesar_form_contacto: {e}")
        return None

# procesar proyecto

def procesar_form_proyecto(dataForm):
    try:
        # Conexión a la base de datos y ejecución de la consulta
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # Consulta SQL para insertar el proyecto en la tabla tbl_proyectos
                sql = """
                    INSERT INTO tbl_proyectos 
                    (nombre_proyecto, descripcion, tipo_evento, fecha_inicio, fecha_fin, id_cliente, id_miembro_responsable) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
                
                # Obtener los valores del formulario
                valores = (
                    dataForm['nombre_proyecto'],
                    dataForm['descripcion'],
                    int(dataForm['tipo_evento']),
                    dataForm['fecha_inicio'],
                    dataForm['fecha_fin'],
                    int(dataForm['id_cliente']),
                    int(dataForm['id_miembro_responsable'])
                )
                
                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()
                
                # Retornar éxito si se inserta correctamente
                return cursor.lastrowid

    except Exception as e:
        print(f"Error en procesar_form_proyecto: {e}")
        return None

# procesar evento

def procesar_form_evento(dataForm):
    try:
        # Conexión a la base de datos y ejecución de la consulta
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                sql = """
                    INSERT INTO tbl_eventos 
                    (nombre_evento, descripcion, fecha_inicio, fecha_termino, ubicacion, id_cliente, id_miembro_responsable, asistencia, foto_empleado)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                
                # Obtener valores del formulario
                valores = (
                    dataForm['nombre_evento'],
                    dataForm['descripcion'],
                    dataForm['fecha_inicio'],
                    dataForm['fecha_termino'],
                    dataForm['ubicacion'],
                    int(dataForm['id_cliente']),
                    int(dataForm['id_miembro_responsable']),
                    int(dataForm['asistencia']),
                    dataForm['foto_empleado']
                )

                cursor.execute(sql, valores)
                conexion_MySQLdb.commit()

                # Retornar éxito si se inserta correctamente
                return cursor.rowcount > 0

    except Exception as e:
        print(f"Error en procesar_form_evento: {e}")
        return None


